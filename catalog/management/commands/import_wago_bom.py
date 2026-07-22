from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date
from decimal import Decimal, InvalidOperation
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.db.models import Max
from openpyxl import load_workbook

from catalog.models import Component, PriceListEntry
from catalog.dimensions import parse_whd_from_name
from catalog.manufacturer import normalize_manufacturer

DATA_START_ROW = 8
DATA_END_ROW = 93
IMPORT_ORDER_DATE = date(2026, 7, 22)


@dataclass
class BomRow:
    sheet_row: int
    serial_number: str
    serial_is_generated: bool
    part_number: str
    name: str
    description: str
    manufacturer: str
    type: str
    width_mm: Decimal | None
    height_mm: Decimal | None
    depth_mm: Decimal | None
    consumed_dc_current_ma: Decimal | None
    quantity: int | None
    price: Decimal | None
    currency: str | None
    flags: list[str] = field(default_factory=list)


def _to_decimal(value) -> Decimal | None:
    if value is None or value == "":
        return None
    try:
        dec = Decimal(str(value))
    except (InvalidOperation, ValueError):
        return None
    if dec == 0:
        return None
    return dec


def _to_dimension(value) -> Decimal | None:
    dec = _to_decimal(value)
    return None if dec is None else dec.quantize(Decimal("0.01"))


def _to_current(value) -> Decimal | None:
    dec = _to_decimal(value)
    return None if dec is None else dec.quantize(Decimal("0.01"))


def _pick_price(ws, row: int) -> tuple[Decimal | None, str | None]:
    checks = [
        (55, "USD"),  # BC USD Local
        (54, "USD"),  # BB USD Foreign
        (56, "EUR"),  # BD EURO Foreign
        (57, "EUR"),  # BE Euro Local
    ]
    for col, currency in checks:
        price = _to_decimal(ws.cell(row, col).value)
        if price is not None:
            return price.quantize(Decimal("0.01")), currency
    return None, None


def _is_section_header(ws, row: int) -> bool:
    item = ws.cell(row, 1).value
    part = ws.cell(row, 7).value
    manufacturer = ws.cell(row, 3).value
    price, _ = _pick_price(ws, row)
    qty_raw = ws.cell(row, 52).value  # AZ
    qty = None
    if qty_raw not in (None, ""):
        try:
            qty = float(qty_raw)
        except (TypeError, ValueError):
            qty = None
    return bool(item) and not part and not manufacturer and price is None and qty is None


def _make_unique_serial(base: str, row: int, used: set[str]) -> str:
    if base not in used:
        return base
    candidate = f"{base}~{row}"
    if candidate not in used:
        return candidate
    n = 2
    while f"{candidate}~{n}" in used:
        n += 1
    return f"{candidate}~{n}"


def parse_bom_workbook(path: Path) -> list[BomRow]:
    wb = load_workbook(path, data_only=True)
    if "Panel BOM" not in wb.sheetnames:
        raise CommandError(f'Sheet "Panel BOM" not found in {path}')
    ws = wb["Panel BOM"]

    current_type = ""
    used_serials: set[str] = set()
    rows: list[BomRow] = []

    for row in range(DATA_START_ROW, DATA_END_ROW + 1):
        if _is_section_header(ws, row):
            current_type = str(ws.cell(row, 1).value).strip()
            continue

        item_raw = ws.cell(row, 1).value
        part_raw = ws.cell(row, 7).value
        item = str(item_raw).strip() if item_raw else ""
        part = str(part_raw).strip() if part_raw else ""
        manufacturer = normalize_manufacturer(str(ws.cell(row, 3).value or ""))
        price, currency = _pick_price(ws, row)

        qty_raw = ws.cell(row, 52).value
        quantity = None
        flags: list[str] = []
        if qty_raw not in (None, ""):
            try:
                qty_float = float(qty_raw)
                if qty_float != int(qty_float):
                    flags.append("fractional_qty")
                quantity = max(0, int(round(qty_float)))
            except (TypeError, ValueError):
                flags.append("invalid_qty")

        if not part and not item:
            continue
        if not part and price is None and quantity is None:
            continue

        serial_is_generated = not bool(part)
        if part:
            serial_number = _make_unique_serial(part, row, used_serials)
            if serial_number != part:
                flags.append("duplicate_part")
        else:
            serial_number = f"BOM-R{row}"
            flags.append("generated_serial")

        used_serials.add(serial_number)

        if price is None:
            flags.append("no_price")

        width_mm = _to_dimension(ws.cell(row, 4).value)
        height_mm = _to_dimension(ws.cell(row, 5).value)
        depth_mm = None
        name_w, name_h, name_d = parse_whd_from_name(item)
        if width_mm is None and name_w is not None:
            width_mm = name_w
        if height_mm is None and name_h is not None:
            height_mm = name_h
        if name_d is not None:
            depth_mm = name_d

        rows.append(
            BomRow(
                sheet_row=row,
                serial_number=serial_number,
                serial_is_generated=serial_is_generated,
                part_number=part,
                name=item,
                description=item,
                manufacturer=manufacturer,
                type=current_type or "Uncategorized",
                width_mm=width_mm,
                height_mm=height_mm,
                depth_mm=depth_mm,
                consumed_dc_current_ma=_to_current(ws.cell(row, 6).value),
                quantity=quantity,
                price=price,
                currency=currency,
                flags=flags,
            )
        )

    return rows


class Command(BaseCommand):
    help = "Import Wago BOM components and price entries from an Excel workbook."

    def add_arguments(self, parser):
        parser.add_argument("xlsx_path", type=str, help="Path to the BOM .xlsx file")
        parser.add_argument(
            "--commit",
            action="store_true",
            help="Write to the database (default is preview only)",
        )

    def handle(self, *args, **options):
        path = Path(options["xlsx_path"])
        if not path.exists():
            raise CommandError(f"File not found: {path}")

        rows = parse_bom_workbook(path)
        if not rows:
            raise CommandError("No importable rows found.")

        if options["commit"]:
            created_components, created_prices = self._commit(rows)
            self.stdout.write(
                self.style.SUCCESS(
                    f"Imported {created_components} components and {created_prices} price entries."
                )
            )
            return

        self._print_preview(rows)

    def _print_preview(self, rows: list[BomRow]) -> None:
        usd = sum(1 for r in rows if r.currency == "USD")
        eur = sum(1 for r in rows if r.currency == "EUR")
        generated = sum(1 for r in rows if r.serial_is_generated)
        no_price = sum(1 for r in rows if r.price is None)

        self.stdout.write(f"Preview: {len(rows)} components")
        self.stdout.write(
            f"  EUR: {eur} | USD: {usd} | generated serials: {generated} | no price: {no_price}"
        )
        self.stdout.write("")
        header = (
            f"{'row':>3}  {'serial':<22} {'gen':>3}  {'mfr':<18} {'qty':>6}  "
            f"{'price':>10} {'cur':>3}  {'flags':<20} name"
        )
        self.stdout.write(header)
        self.stdout.write("-" * len(header))

        for r in rows:
            flags = ",".join(r.flags) if r.flags else "-"
            qty = str(r.quantity) if r.quantity is not None else "-"
            price = f"{r.price:.2f}" if r.price is not None else "-"
            cur = r.currency or "-"
            name = r.name[:48] + ("…" if len(r.name) > 48 else "")
            self.stdout.write(
                f"{r.sheet_row:>3}  {r.serial_number:<22} "
                f"{'Y' if r.serial_is_generated else 'N':>3}  "
                f"{r.manufacturer[:18]:<18} {qty:>6}  {price:>10} {cur:>3}  "
                f"{flags:<20} {name}"
            )

        self.stdout.write("")
        self.stdout.write("No database changes made. Re-run with --commit to import.")

    @transaction.atomic
    def _commit(self, rows: list[BomRow]) -> tuple[int, int]:
        next_pricelist = (Component.objects.aggregate(m=Max("pricelist_id"))["m"] or 0) + 1
        created_components = 0
        created_prices = 0

        for row in rows:
            if Component.objects.filter(serial_number=row.serial_number).exists():
                raise CommandError(
                    f"Component {row.serial_number} already exists; aborting import."
                )

            component = Component.objects.create(
                serial_number=row.serial_number,
                name=row.name,
                description=row.description,
                type=row.type,
                part_number=row.part_number,
                manufacturer=row.manufacturer,
                pricelist_id=next_pricelist,
                width_mm=row.width_mm,
                height_mm=row.height_mm,
                depth_mm=row.depth_mm,
                consumed_dc_current_ma=row.consumed_dc_current_ma,
                serial_is_generated=row.serial_is_generated,
            )
            next_pricelist += 1
            created_components += 1

            if row.price is not None and row.quantity is not None and row.currency:
                PriceListEntry.objects.create(
                    pricelist_id=component.pricelist_id,
                    component=component,
                    price=row.price,
                    quantity=row.quantity,
                    currency=row.currency,
                    order_time=IMPORT_ORDER_DATE,
                )
                created_prices += 1

        return created_components, created_prices
